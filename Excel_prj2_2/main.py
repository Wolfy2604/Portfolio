import os
import xlrd
import openpyxl
from pprint import pprint

import rel_district as rarea
import rel_disease as rdis

order_lst = ['Ale', 'Bch', 'Bez', 'Bgl', 'Bog', 'Bor', 'Cha', 'ChV', 'Elx', 'Gig', 'Isa', 'Kam', 'Kch',
             'Kin', 'Kla', 'Kos', 'Kra', 'Nef', 'Nov', 'Okt', 'Otr', 'Pes', 'Poh', 'Pri', 'Sam', 'Ser', 'Jdr',
             'She', 'Shi', 'Sta', 'SyR', 'Syz', 'Tlt', 'Vol', 'Xvo', 'Yar']

order_lst.reverse()


def take_data():
    book = os.listdir(path='./Ввод')[0]
    sheet = openpyxl.load_workbook(f'./Ввод/{book}').active
    total_dict = {}
    for x in range(107):
        total_dict.update({x: {}})


    # area_idx = rarea.relation_list[reg_book[:3]]
    for row in sheet.values:
        print(row[0])
        if not row[0]:
            continue
        elif row[0].find('Территория') != -1:
            try:
                print('LIST POP')
                area = order_lst.pop()
            except IndexError:
                break
            try:
                area_idx = rarea.relation_list[area]
                print(f'{area} - {area_idx}')
            except KeyError:
                continue

        try:
            dis_idx = int(row[0])
            row_val = []
            for cell in row[4:]:
                row_val.append(cell)
            print(f'ROW_VAL {row_val}')
            total_dict[dis_idx].update({area_idx: row_val})
        except ValueError:
            continue

    pprint(total_dict)
    write_data(total_dict)


def write_data(data):
    exit_file = os.listdir(path='./Вывод')[0]
    wb = openpyxl.load_workbook(f'./Вывод/{exit_file}')
    ws = wb.active

    for key_dis, val in data.items():
        key_dis = str(key_dis)
        if str(key_dis) in rdis.relation_list.keys():
            dis_row = rdis.relation_list[key_dis]
            for key_area, row in val.items():
                local = dis_row + key_area - 1
                # print(f'NEW AREA {key_area}')
                # print(f'START {local}')
                # print(f'VAL {row}')
                start_col = 3
                for cell_val in row:
                    ws.cell(row=local, column=start_col, value=cell_val)
                    start_col += 1
    wb.save(filename='./Вывод/1.xlsx')


take_data()